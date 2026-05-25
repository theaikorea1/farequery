"""
KITA 해상·항공 참고운임 수집기 - HTML POST 방식 (고속버전)
===========================================================
URL: https://kita.net/shippers/logisticsFare/logisticsSeaFare.do

playwright 방식 대비 10배 빠름 (도시당 1~3초 vs 15~30초)
2015년 전체 수집: 약 3시간 (playwright는 30시간+)

원리:
  1. playwright로 세션 쿠키 1회 획득 (약 30초)
  2. requests.post()로 HTML 직접 수신
  3. BeautifulSoup으로 테이블 파싱
  4. 도시 코드 목록: logisticsSeaCodeView.do / logisticsAirCodeView.do

실행:
  python kita_api_scraper.py              ← 전체 (2015년~현재)
  python kita_api_scraper.py --from 2025  ← 테스트 (2025년~현재)
  python kita_api_scraper.py --retry      ← 실패 항목만 재수집
  python kita_api_scraper.py --schedule   ← 매월 9일 08:00 자동
  Ctrl+C                                  ← 안전 종료

결과:
  API_output/kita_all_freight.csv
  API_output/kita_sea_freight.csv
  API_output/kita_air_freight.csv
  API_output/kita_collection_summary.xlsx  ← 수집 현황 Summary (4개 시트)
  API_output/progress/               ← 도시별 중간 저장
  API_output/failed_cities.json      ← 실패 도시 목록
"""

import re, csv, json, time, logging, sys, signal
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ── 로그 ──────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("kita_api_scraper.log", encoding="utf-8", mode="a"),
    ],
)
log = logging.getLogger(__name__)

# ── 설정 ──────────────────────────────────────────────────────
DEFAULT_START_YEAR  = 2015
DEFAULT_START_MONTH = 1
SCRIPT_DIR   = Path(__file__).resolve().parent
REPO_ROOT    = SCRIPT_DIR.parent
OUTPUT_DIR   = REPO_ROOT / "FareDB"
PROGRESS_DIR = SCRIPT_DIR / "progress"
FAILED_FILE  = SCRIPT_DIR / "failed_cities.json"
SUMMARY_FILE = OUTPUT_DIR / "kita_collection_summary.xlsx"

COLUMNS = [
    "분류", "출발(도시)", "도착(국가)", "도착(도시)",
    "년", "월",
    "TEU", "FEU", "단위(USD-해상)",
    "100kg", "300kg", "500kg", "단위(USD-항공)",
    "업데이트 일자",
]
META_SEA = (
    "# [해상] 출처: 한국무역협회(KITA) | "
    "시장 평균 해상운임(Ocean Freight), 부산발 | "
    "부대비용 별도 / 이용 선사·물동량·결제조건에 따라 상이 | "
    "운임 업로드: 매월 첫째 주 | 단위: USD"
)
META_AIR = (
    "# [항공] 출처: 한국무역협회(KITA) | "
    "시장 평균 항공운임(Air Freight), 부산발 | "
    "부대비용 별도 / 이용 항공사·물동량·결제조건에 따라 상이 | "
    "공시 운임 구성: 기본 항공운임 + 유류할증료 + 터미널 화물처리비용 | "
    "운임 업로드: 매월 첫째 주 | 단위: USD"
)

# 해상/항공 URL 및 코드 설정
SEA_PAGE    = "https://kita.net/shippers/logisticsFare/logisticsSeaFare.do"
AIR_PAGE    = "https://kita.net/shippers/logisticsFare/logisticsAirFare.do"
SEA_CODE_API= "https://kita.net/shippers/logisticsFare/logisticsSeaCodeView.do"
AIR_CODE_API= "https://kita.net/shippers/logisticsFare/logisticsAirCodeView.do"

# 해상 권역코드
SEA_REGIONS = {
    "010101":"북미",     "010102":"중남미",   "010103":"유럽",
    "010104":"아시아",   "010105":"일본",     "010106":"중국",
    "010107":"아프리카", "010108":"오세아니아","010109":"중동",
    "010110":"러시아/CIS",
}
# 항공 권역코드 (02로 시작)
AIR_REGIONS = {
    "020101":"북미",     "020102":"중남미",   "020103":"유럽",
    "020104":"아시아",   "020105":"일본",     "020106":"중국",
    "020107":"아프리카", "020108":"오세아니아","020109":"중동",
    "020110":"러시아",
}

SEA_START = "0102010102"   # 부산
AIR_START = "0202010101"   # 인천


# ── 숫자 클린업 ───────────────────────────────────────────────
def clean_num_from_td(td_tag, field_name: str = "") -> str:
    """
    BeautifulSoup td 태그에서 운임값만 정확히 추출.

    KITA HTML 구조 (이미지 확인):
      정상 (운임+편차):  <td>9130<span class="up_num">▲9130</span></td>
                         → 운임=9130  (span 제거 후 직접 텍스트)
      운임없음(편차만):  <td><span class="dn_num">▼-8740</span></td>
                         → 운임=''    (span 제거하면 텍스트 없음 → 빈값)
      운임+편차:         <td>1400<span class="dn_num">▼-500</span></td>
                         → 운임=1400  (span 제거 후 직접 텍스트)

    핵심: span(편차 영역) 제거 후 직접 텍스트노드만 읽음
    → 운임값 없는 달은 자동으로 '' 반환 (음수 오파싱 원천 차단)
    """
    if td_tag is None:
        return ""

    # span 태그(편차) 제거 → 운임값 텍스트만 남김
    import copy
    td = copy.copy(td_tag)
    for span in td.find_all(["span", "button", "a"]):
        span.decompose()

    raw = td.get_text(strip=True).replace(",", "")

    if not raw or raw == "-":
        return ""

    try:
        val = float(raw)
        result = str(round(val, 1))
        if val < 0:
            log.warning(
                f"  ⚠️  음수 운임값 [{field_name}]: {result} "
                f"(원본HTML: {str(td_tag)[:60]}) → KITA 원본 확인 필요"
            )
        return result
    except ValueError:
        return ""


def clean_num(raw: str, field_name: str = "") -> str:
    """문자열 fallback용 (td 태그 없을 때)"""
    if not raw:
        return ""
    raw = re.sub(r"[↑↓▲▼▶◀]", "", raw)
    raw = re.sub(r"\(.*?\)", "", raw)
    raw = raw.replace(",", "").strip()
    if not raw or raw == "-":
        return ""
    try:
        return str(round(float(raw.split()[0]), 1))
    except (ValueError, IndexError):
        return ""


# ── 중간 저장 / 실패 기록 ─────────────────────────────────────
def save_progress(rows, freight_type, region_code, city_code):
    PROGRESS_DIR.mkdir(parents=True, exist_ok=True)
    ts    = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"{freight_type}_{region_code}_{city_code}_{ts}.csv"
    with open(PROGRESS_DIR / fname, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        for row in rows:
            w.writerow({col: row.get(col, "") for col in COLUMNS})
    log.info(f"    💾 중간 저장: {fname} ({len(rows)}행)")


def load_failed():
    if not FAILED_FILE.exists(): return []
    try: return json.loads(FAILED_FILE.read_text(encoding="utf-8"))
    except: return []

def append_failed(entry):
    OUTPUT_DIR.mkdir(exist_ok=True)
    existing = [e for e in load_failed()
                if not (e["freight_type"]==entry["freight_type"]
                        and e["city_code"]==entry["city_code"])]
    existing.append(entry)
    FAILED_FILE.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")

def remove_failed(freight_type, city_code):
    existing = [e for e in load_failed()
                if not (e["freight_type"]==freight_type and e["city_code"]==city_code)]
    if FAILED_FILE.exists():
        FAILED_FILE.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")

def get_done_cities():
    if not PROGRESS_DIR.exists(): return set()
    done = set()
    for f in PROGRESS_DIR.glob("*.csv"):
        parts = f.stem.split("_")
        if len(parts) >= 3:
            done.add((parts[0], parts[2]))
    return done

def _report_missing_months(rows: list, start_year: int, start_month: int,
                             freight_type: str):
    """
    수집된 데이터에서 누락 월을 탐지하여 로그 출력 및 JSON 저장.
    원본 KITA 미집계 vs 수집 오류 구분 근거 자료로 활용.
    데이터는 원본 그대로 보존 (누락 월 행 삽입 안 함).
    """
    if not rows:
        return

    from collections import defaultdict

    # 수집된 (년, 월) 집합 — 도시 상관없이 전체 기준
    collected = set((r["년"], r["월"]) for r in rows)

    # 기대 범위: start_year/month ~ 현재 월
    today = date.today()
    expected = set()
    y, m = start_year, start_month
    while (y, int(m)) <= (today.year, today.month):
        expected.add((str(y), f"{m:02d}"))
        m += 1
        if m > 12:
            m = 1; y += 1

    missing = sorted(expected - collected)

    if missing:
        sep = "=" * 50
        log.warning(f"\n{sep}")
        log.warning(f"[{freight_type}] 누락 월 {len(missing)}개 (KITA 원본 미집계 가능 — Summary xlsx 참조)")
        for ym in missing:
            log.warning(f"   {ym[0]}년 {ym[1]}월")
        log.warning(sep)
        # ※ 누락 월 상세는 Summary xlsx '운임값미배정(비표시)' 시트에 기록됨
        #   → missing_months_*.json 별도 파일 생성 안 함
    else:
        log.info(f"[{freight_type}] 월 누락 없음 ✅")
    return missing


def merge_progress_to_csv():
    if not PROGRESS_DIR.exists():
        log.warning("progress 폴더 없음")
        return
    files = sorted(PROGRESS_DIR.glob("*.csv"))
    latest = {}
    for f in files:
        parts = f.stem.split("_")
        if len(parts) >= 3:
            latest[f"{parts[0]}_{parts[1]}_{parts[2]}"] = f

    all_rows = []
    for f in latest.values():
        try:
            with open(f, encoding="utf-8-sig", newline="") as fp:
                all_rows.extend(list(csv.DictReader(fp)))
        except Exception as e:
            log.error(f"파일 읽기 실패: {f} - {e}")

    # [버그 수정] 깨진 행 제거: 년/월이 없거나 출발도시가 비정상인 행
    def _is_valid_row(r):
        try:
            y = str(r.get("년","")).strip()
            m = str(r.get("월","")).strip()
            dep = str(r.get("출발(도시)","")).strip()
            if not y or not m or not y.isdigit() or not m.isdigit():
                return False
            if int(y) < 2000 or int(y) > 2100:
                return False
            if int(m) < 1 or int(m) > 12:
                return False
            # 출발도시 정상값 검증
            if dep and dep not in ("Busan", "Incheon"):
                return False
            return True
        except:
            return False

    before = len(all_rows)
    all_rows = [r for r in all_rows if _is_valid_row(r)]
    if before != len(all_rows):
        log.warning(f"  ⚠️  깨진 행 {before - len(all_rows)}개 제거 (HTML 파싱 잔재)")

    sea = [r for r in all_rows if r.get("분류") == "해상"]
    air = [r for r in all_rows if r.get("분류") == "항공"]
    OUTPUT_DIR.mkdir(exist_ok=True)

    # 최신본 (항상 덮어쓰기)
    save_csv(all_rows, OUTPUT_DIR / "kita_all_freight.csv", [META_SEA, META_AIR])
    save_csv(sea,      OUTPUT_DIR / "kita_sea_freight.csv", [META_SEA])
    save_csv(air,      OUTPUT_DIR / "kita_air_freight.csv", [META_AIR])

    # 수집일자 백업본 — kita_all_freight.csv 는 별도 파일로도 유지
    # Summary xlsx 전체데이터 시트와 별개로 CSV 원본 파일을 보존
    ts_date = datetime.now().strftime("%Y%m%d")
    backup_path = OUTPUT_DIR / f"kita_all_freight_{ts_date}.csv"
    if not backup_path.exists():   # 같은 날 중복 실행 시 덮어쓰지 않음
        save_csv(all_rows, backup_path, [META_SEA, META_AIR])
        log.info(f"백업 저장: {backup_path.name}")

    log.info(f"병합 완료: 해상 {len(sea)}행 / 항공 {len(air)}행 / 전체 {len(all_rows)}행")

    # 누락 월 리포트 (원본 보존, 삽입 안 함)
    # 실제 수집된 데이터의 최소 년/월을 기준으로 비교 (--from 옵션 반영)
    log.info("\n누락 월 검사 중...")
    def _actual_start(rows):
        if not rows: return DEFAULT_START_YEAR, DEFAULT_START_MONTH
        valid = []
        for r in rows:
            try:
                y = int(r["년"])
                m = int(r["월"])
                if 2000 <= y <= 2100 and 1 <= m <= 12:
                    valid.append((y, m))
            except (ValueError, KeyError, TypeError):
                continue
        if not valid: return DEFAULT_START_YEAR, DEFAULT_START_MONTH
        return sorted(valid)[0]
    sea_sy, sea_sm = _actual_start(sea)
    air_sy, air_sm = _actual_start(air)
    _report_missing_months(sea, sea_sy, sea_sm, "해상")
    _report_missing_months(air, air_sy, air_sm, "항공")

    # Summary xlsx 생성/업데이트
    try:
        _sy = min((int(r["년"]) for r in all_rows), default=DEFAULT_START_YEAR)
        _sm = min((int(r["월"]) for r in all_rows if int(r["년"])==_sy), default=DEFAULT_START_MONTH)
        run_at = generate_summary_xlsx(
            all_rows, _sy, _sm, SUMMARY_FILE,
            all_csv_path=OUTPUT_DIR / "kita_all_freight.csv"
        )
        log.info(f"\nSummary 저장: {SUMMARY_FILE}")
    except Exception as e:
        log.error(f"Summary 생성 실패 (수집 결과에는 영향 없음): {e}")

def save_csv(rows, path, meta_lines=None):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        if meta_lines:
            for line in meta_lines:
                f.write(line + "\n")
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        for row in rows:
            w.writerow({col: row.get(col, "") for col in COLUMNS})
    log.info(f"저장: {path} ({len(rows)}행)")



# ══════════════════════════════════════════════════════════════
# Summary xlsx 생성 (v3)
# ══════════════════════════════════════════════════════════════
"""Summary xlsx 개선 v2 — 모든 요청사항 반영"""
import json, copy
from datetime import datetime, date
from pathlib import Path
from collections import defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

C = {
    "navy":"1B3A6B","blue":"2563EB","sky":"EFF6FF","gray_hd":"E2E8F0",
    "gray_bg":"F5F7FA","white":"FFFFFF","green":"166534","green_bg":"DCFCE7",
    "amber":"92400E","amber_bg":"FEF3C7","red":"991B1B","red_bg":"FEE2E2",
    "dark":"1E293B","mid":"64748B","light":"94A3B8","orange":"C2410C","orange_bg":"FFF7ED",
}
SEA_URL = "https://kita.net/shippers/logisticsFare/logisticsSeaFare.do"
AIR_URL = "https://kita.net/shippers/logisticsFare/logisticsAirFare.do"

def _fill(c): return PatternFill("solid", fgColor=c)
def _font(c=C["dark"], bold=False, sz=10, name="맑은 고딕"):
    return Font(name=name, size=sz, bold=bold, color=c)
def _bdr(c=C["gray_hd"]):
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def _aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _c(ws, r, c, v, bold=False, bg=None, fc=C["dark"], align="left",
       sz=10, wrap=False, border=True, merge_to=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = _font(fc, bold, sz)
    cell.alignment = _aln(align, "center", wrap)
    if bg: cell.fill = _fill(bg)
    if border: cell.border = _bdr()
    if merge_to:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=merge_to)
    return cell
def _hdr(ws, row, labels, c_start=1, bg=C["navy"], fc=C["white"], h=22):
    for i, lbl in enumerate(labels, c_start):
        _c(ws, row, i, lbl, bold=True, bg=bg, fc=fc, align="center")
    ws.row_dimensions[row].height = h

def _get_missing_detail(rows, sy, sm, freight_type):
    """누락 분석: 전체 누락 월 + 특정 도시 누락 월 반환"""
    today = date.today()
    # 수집된 (년,월) 전체 집합
    collected_ym = set((str(int(r["년"])), f"{int(r['월']):02d}") for r in rows)
    # 기대 전체 월
    expected = set()
    y, m = sy, sm
    while (y, m) <= (today.year, today.month):
        expected.add((str(y), f"{m:02d}")); m += 1
        if m > 12: m=1; y+=1
    # 전체 누락 월 (전 도시 없음)
    full_missing = sorted(expected - collected_ym)
    # 도시별 누락 월 (해당 년/월은 있지만 특정 도시 없음)
    all_cities = set((r.get("출발(도시)",""), r.get("도착(국가)",""), r.get("도착(도시)","")) for r in rows)
    city_missing = defaultdict(list)  # (출발,국가,도시) → 누락월 리스트
    for ym in sorted(expected - set(full_missing)):
        y_val, m_val = ym
        present = set(
            (r.get("출발(도시)",""), r.get("도착(국가)",""), r.get("도착(도시)",""))
            for r in rows if str(int(r["년"]))==y_val and f"{int(r['월']):02d}"==m_val
        )
        for city in sorted(all_cities - present):
            city_missing[city].append(f"{y_val}-{m_val}")
    return full_missing, city_missing

def _get_empty_detail(rows, freight_type):
    """운임값 공백 행: 년/월은 있지만 운임값이 없는 경우"""
    empty = []
    for r in rows:
        if freight_type == "sea":
            if not r.get("TEU") and not r.get("FEU"):
                empty.append(r)
        else:
            if not r.get("100kg") and not r.get("300kg") and not r.get("500kg"):
                empty.append(r)
    return empty


# ─────────────────────────────────────────────────────────────
# Sheet 1: 개요(Overview)
# ─────────────────────────────────────────────────────────────
def _overview(ws, all_rows, sy, sm, run_at):
    ws.title = "개요(Overview)"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 62

    sea = [r for r in all_rows if r.get("분류")=="해상"]
    air = [r for r in all_rows if r.get("분류")=="항공"]

    # 타이틀
    ws.merge_cells("A1:B1")
    _c(ws,1,1,"KITA 해상·항공 참고운임 수집 Summary",
       bold=True,bg=C["navy"],fc=C["white"],align="center",sz=14,border=False)
    ws.row_dimensions[1].height = 32

    def section(row, txt):
        ws.merge_cells(f"A{row}:B{row}")
        _c(ws,row,1,f"  {txt}",bold=True,bg=C["gray_hd"],fc=C["navy"],align="left",border=False)
        ws.row_dimensions[row].height = 20

    def kv(row, k, v, vbg=None, vfc=C["dark"], wrap=False):
        _c(ws,row,1,k,bold=True,bg=C["gray_bg"],fc=C["mid"])
        _c(ws,row,2,v,bg=vbg or C["white"],fc=vfc,wrap=wrap)
        ws.row_dimensions[row].height = 20 if not wrap else 32

    section(2, "▸ 수집 설정")
    kv(3,"해상 수집 URL", SEA_URL)
    kv(4,"항공 수집 URL", AIR_URL)
    kv(5,"수집 방식","HTML POST (requests) + BeautifulSoup 파싱 / 세션: playwright 1회")
    kv(6,"정기 수집 스케줄","매월 9일 08:00 KST  (KITA 업로드: 매월 첫째 주)")

    section(8, "▸ 이번 수집 결과")
    kv(9,"수집 실행 일시", run_at)

    # 기간 + 개월 수
    def period_str(rows):
        if not rows: return "-", 0
        yms = sorted(set((str(int(r["년"])),f"{int(r['월']):02d}") for r in rows))
        start = f"{yms[0][0]}년 {yms[0][1]}월"
        end   = f"{yms[-1][0]}년 {yms[-1][1]}월"
        return f"{start}  ~  {end}", len(yms)

    sea_period, sea_months = period_str(sea)
    air_period, air_months = period_str(air)
    kv(10,"해상 수집 기간", f"{sea_period}  ({sea_months}개월 수집)", vfc=C["navy"])
    kv(11,"항공 수집 기간", f"{air_period}  ({air_months}개월 수집)", vfc=C["navy"])
    kv(12,"해상 수집 행수", len(sea), vfc=C["navy"])
    kv(13,"항공 수집 행수", len(air), vfc=C["navy"])
    kv(14,"전체 수집 행수", len(all_rows), vfc=C["navy"])

    sea_nations = len(set(r["도착(국가)"] for r in sea))
    sea_cities  = len(set((r["도착(국가)"],r["도착(도시)"]) for r in sea))
    air_nations = len(set(r["도착(국가)"] for r in air))
    air_cities  = len(set((r["도착(국가)"],r["도착(도시)"]) for r in air))
    kv(15,"해상 수집 권역 수", f"{sea_nations}개 권역")
    kv(16,"해상 수집 도시 수", f"{sea_cities}개 도시")
    kv(17,"항공 수집 권역 수", f"{air_nations}개 권역")
    kv(18,"항공 수집 도시 수", f"{air_cities}개 도시")

    # 운임값 미배정(비표시) 안내
    section(20, "▸ 운임값 미배정(비표시)  ← 시트 '운임값미배정(비표시)' 참조")
    ws.merge_cells("A21:B21")
    desc = (
        "특정 출발지-도착지 구간에서 해당 월 KITA가 운임을 집계·배정하지 않은 경우입니다. "
        "수집 오류가 아닌 KITA 원본 미집계이므로 데이터 품질에 문제는 없습니다. "
        "상세 내역은 '운임값미배정(비표시)' 시트를 참조하세요."
    )
    _c(ws,21,1,desc,bg=C["orange_bg"],fc=C["orange"],wrap=True,sz=10)
    ws.row_dimensions[21].height = 48

    sea_full, sea_city_miss = _get_missing_detail(sea, sy, sm, "sea")
    air_full, air_city_miss = _get_missing_detail(air, sy, sm, "air")
    sea_empty = _get_empty_detail(sea, "sea")
    air_empty = _get_empty_detail(air, "air")

    kv(22,"해상 비표시 현황",
       f"전체 누락 월: {len(sea_full)}개월  /  특정 구간 누락: {sum(len(v) for v in sea_city_miss.values())}건  /  운임값 공백: {len(sea_empty)}건")
    kv(23,"항공 비표시 현황",
       f"전체 누락 월: {len(air_full)}개월  /  특정 구간 누락: {sum(len(v) for v in air_city_miss.values())}건  /  운임값 공백: {len(air_empty)}건")

    # 운임값 미배정(공백) 안내
    section(25, "▸ 운임값 미배정(공백)  ← 시트 '운임값미배정(공백)' 참조")
    ws.merge_cells("A26:B26")
    desc2 = (
        "해당 년/월 데이터 행은 존재하지만 TEU/FEU(해상) 또는 100kg/300kg/500kg(항공) "
        "값이 공백인 경우입니다. KITA가 해당 구간의 운임을 미배정한 것이며 수집 오류가 아닙니다. "
        "상세 내역은 '운임값미배정(공백)' 시트를 참조하세요."
    )
    _c(ws,26,1,desc2,bg=C["amber_bg"],fc=C["amber"],wrap=True,sz=10)
    ws.row_dimensions[26].height = 48


# ─────────────────────────────────────────────────────────────
# Sheet 2: 도시별이력(City_Log) — 출발도시를 앞에 배치
# ─────────────────────────────────────────────────────────────
def _city_log(ws, all_rows, run_at, is_update=False):
    ws.title = "도시별이력(City_Log)"
    # [수정] 출발(도시)를 도착(국가) 앞에 배치
    FIXED = ["분류", "출발(도시)", "도착(국가)", "도착(도시)"]
    FIXED_W = [8, 12, 14, 18]

    existing_data = {}
    max_hist_col = len(FIXED)

    if is_update and ws.max_row > 1:
        for row in ws.iter_rows(min_row=2, values_only=True):
            key = tuple(row[:4])
            existing_data[key] = list(row)
        max_hist_col = ws.max_column
    else:
        ws.delete_rows(1, ws.max_row)

    city_stats = defaultdict(int)
    for r in all_rows:
        key = (r.get("분류",""), r.get("출발(도시)",""),
               r.get("도착(국가)",""), r.get("도착(도시)",""))
        city_stats[key] += 1

    run_dt = run_at[:16] if run_at else datetime.now().strftime("%Y-%m-%d %H:%M")
    all_keys = set(existing_data.keys()) | set(city_stats.keys())

    prev_hdrs = [ws.cell(1,c).value for c in range(len(FIXED)+1, max_hist_col+1)] \
                if is_update else []
    hdrs = FIXED + prev_hdrs + [f"{run_dt} 행수"]

    _hdr(ws, 1, hdrs)
    for i, w in enumerate(FIXED_W, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in range(len(FIXED)+1, len(hdrs)+1):
        ws.column_dimensions[get_column_letter(c)].width = 22

    sorted_keys = sorted(all_keys, key=lambda x: (x[0], x[2], x[3]))
    for ri, key in enumerate(sorted_keys, 2):
        bg = C["gray_bg"] if ri%2==1 else C["white"]
        for ci, val in enumerate(key, 1):
            _c(ws, ri, ci, val, bg=bg,
               fc=C["navy"] if ci<=3 else C["dark"], bold=(ci<=2))
        if is_update and key in existing_data:
            for ci in range(len(FIXED)+1, max_hist_col+1):
                val = existing_data[key][ci-1] if ci-1<len(existing_data[key]) else ""
                _c(ws, ri, ci, val, bg=bg)
        cnt = city_stats.get(key, 0)
        _c(ws, ri, len(hdrs), cnt if cnt else "-",
           bg=C["sky"] if cnt else C["red_bg"],
           fc=C["blue"] if cnt else C["red"], align="center")
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(hdrs))}1"


# ─────────────────────────────────────────────────────────────
# Sheet 3: 운임값미배정(비표시) — 구간 정보 포함
# ─────────────────────────────────────────────────────────────
def _missing_display(ws, all_rows, sy, sm):
    ws.title = "운임값미배정(비표시)"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 55

    sea = [r for r in all_rows if r.get("분류")=="해상"]
    air = [r for r in all_rows if r.get("분류")=="항공"]

    # 설명 박스
    ws.merge_cells("A1:E1")
    _c(ws,1,1,
       "운임값 미배정(비표시): KITA가 해당 월 또는 특정 출발-도착 구간에 운임을 집계·배정하지 않은 경우. "
       "수집 오류가 아닌 KITA 원본 미집계이므로 데이터 품질 이슈 없음.",
       bg=C["orange_bg"],fc=C["orange"],wrap=True,sz=10,border=False)
    ws.row_dimensions[1].height = 36

    _hdr(ws, 2, ["분류","구분","년-월","출발→도착 구간","비고"])
    ws.row_dimensions[2].height = 22

    ri = 3
    for ft, rows in [("해상", sea), ("항공", air)]:
        full_miss, city_miss = _get_missing_detail(rows, sy, sm, ft)

        # 전체 누락 월 (모든 도시 없음)
        for ym in full_miss:
            y, m = ym[0], ym[1]   # tuple (년, 월)
            _c(ws,ri,1, ft, bg=C["red_bg"], fc=C["red"], align="center")
            _c(ws,ri,2, "전체 누락", bg=C["red_bg"], fc=C["red"], align="center")
            _c(ws,ri,3, f"{y}-{m}", bg=C["red_bg"], fc=C["red"], align="center")
            ws.merge_cells(f"D{ri}:E{ri}")
            _c(ws,ri,4, "해당 월 전체 도시 미집계 (KITA 원본)", bg=C["red_bg"], fc=C["red"])
            ws.row_dimensions[ri].height = 18; ri += 1

        # 특정 구간 누락
        for (dep, nation, city), months in sorted(city_miss.items(), key=lambda x: (x[0][1],x[0][2])):
            months_str = ", ".join(months)
            _c(ws,ri,1, ft, bg=C["orange_bg"], fc=C["orange"], align="center")
            _c(ws,ri,2, "구간 누락", bg=C["orange_bg"], fc=C["orange"], align="center")
            _c(ws,ri,3, months_str, bg=C["orange_bg"], fc=C["orange"], align="center", wrap=True)
            route = f"{dep} → {nation} / {city}"
            _c(ws,ri,4, route, bg=C["orange_bg"], fc=C["orange"])
            _c(ws,ri,5, "해당 구간 해당 월 KITA 미집계", bg=C["orange_bg"], fc=C["orange"])
            ws.row_dimensions[ri].height = 18; ri += 1

    ws.freeze_panes = "A3"


# ─────────────────────────────────────────────────────────────
# Sheet 4: 운임값미배정(공백) — 행 존재하지만 값 없음
# ─────────────────────────────────────────────────────────────
def _empty_values(ws, all_rows):
    ws.title = "운임값미배정(공백)"

    hdrs = ["분류","출발(도시)","도착(국가)","도착(도시)","년","월",
            "TEU","FEU","100kg","300kg","500kg","비고"]
    widths = [8,12,14,18,6,6,9,9,9,9,9,40]
    _hdr(ws, 1, hdrs)
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 설명 행
    ws.merge_cells("A2:L2")
    _c(ws,2,1,
       "운임값 미배정(공백): 해당 년/월 데이터 행은 수집됐으나 운임값(TEU/FEU 또는 100kg 등)이 공백인 경우. "
       "KITA 사이트에서 해당 구간의 운임을 배정하지 않은 것이며, span(편차)만 표시된 td에서 운임값 없음을 정확히 감지한 결과.",
       bg=C["amber_bg"],fc=C["amber"],wrap=True,sz=10,border=False)
    ws.row_dimensions[2].height = 48

    sea = [r for r in all_rows if r.get("분류")=="해상"]
    air = [r for r in all_rows if r.get("분류")=="항공"]
    empty_sea = _get_empty_detail(sea, "sea")
    empty_air = _get_empty_detail(air, "air")
    issues = sorted(empty_sea + empty_air,
                    key=lambda r: (r["분류"],r["도착(국가)"],r["도착(도시)"],r["년"],r["월"]))

    ri = 3
    if not issues:
        ws.merge_cells("A3:L3")
        _c(ws,3,1,"운임값 공백 행 없음 ✅",align="center",bg=C["green_bg"],fc=C["green"],bold=True)
    else:
        for r in issues:
            row_data = [r.get("분류",""),r.get("출발(도시)",""),r.get("도착(국가)",""),
                        r.get("도착(도시)",""),r.get("년",""),r.get("월",""),
                        r.get("TEU",""),r.get("FEU",""),
                        r.get("100kg",""),r.get("300kg",""),r.get("500kg",""),
                        "KITA 운임 미배정 (정상)"]
            for ci,val in enumerate(row_data,1):
                _c(ws,ri,ci,val,
                   bg=C["amber_bg"] if ci==12 else C["white"],
                   fc=C["amber"] if ci==12 else C["dark"],
                   align="center" if ci in (1,2,3,4,5,6) else "left")
            ws.row_dimensions[ri].height = 18; ri += 1

    ws.freeze_panes = "A3"
    if ri > 3:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(hdrs))}1"



# ─────────────────────────────────────────────────────────────
# Sheet 5: 전체데이터 — kita_all_freight.csv 내용 그대로
# ─────────────────────────────────────────────────────────────
def _all_data(ws, all_rows):
    ws.title = "전체데이터"

    COLS = ["분류","출발(도시)","도착(국가)","도착(도시)","년","월",
            "TEU","FEU","단위(USD-해상)","100kg","300kg","500kg","단위(USD-항공)","업데이트 일자"]
    COL_W = [8,12,14,18,6,6,10,10,14,10,10,10,14,16]

    # 설명 행
    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    _c(ws,1,1,
       f"kita_all_freight.csv 전체 데이터  |  해상+항공 통합  |  총 {len(all_rows)}행  "
       f"|  생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
       bg=C["navy"], fc=C["white"], bold=True, align="center", sz=10, border=False)
    ws.row_dimensions[1].height = 22

    _hdr(ws, 2, COLS, bg=C["navy"])
    for i, w in enumerate(COL_W, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 22

    # 숫자 컬럼
    num_cols = {"TEU","FEU","100kg","300kg","500kg"}
    ctr_cols = {"분류","년","월","단위(USD-해상)","단위(USD-항공)","업데이트 일자"}

    sorted_rows = sorted(all_rows,
        key=lambda r: (r.get("분류",""), r.get("도착(국가)",""),
                       r.get("도착(도시)",""), r.get("년",""), r.get("월","")))

    for ri, row in enumerate(sorted_rows, 3):
        bg = C["gray_bg"] if ri % 2 == 0 else C["white"]
        ft = row.get("분류","")
        for ci, col in enumerate(COLS, 1):
            val = row.get(col, "")
            if col in num_cols and val not in ("", None):
                try: val = float(val)
                except: pass
            c = ws.cell(row=ri, column=ci, value=val)
            c.font      = _font(C["dark"] if ft=="해상" else C["green"] if ft=="항공" else C["dark"],
                                sz=9)
            c.alignment = _aln("center" if col in ctr_cols else "left")
            c.fill      = _fill(bg)
            c.border    = _bdr(C["gray_hd"])
        ws.row_dimensions[ri].height = 16

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLS))}2"


# ─────────────────────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────────────────────
def generate_summary_xlsx(all_rows, start_year, start_month, summary_path, all_csv_path=None):
    run_at   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    is_upd   = summary_path.exists()

    if is_upd:
        wb = load_workbook(summary_path)
        for sn in list(wb.sheetnames):
            if sn != "도시별이력(City_Log)":
                del wb[sn]
        ws_city = wb["도시별이력(City_Log)"] if "도시별이력(City_Log)" in wb.sheetnames \
                  else wb.create_sheet("도시별이력(City_Log)")
    else:
        wb = Workbook(); wb.remove(wb.active)
        ws_city = wb.create_sheet("도시별이력(City_Log)")

    ws_ov  = wb.create_sheet("개요(Overview)", 0)
    ws_md  = wb.create_sheet("운임값미배정(비표시)")
    ws_ev  = wb.create_sheet("운임값미배정(공백)")
    ws_all = wb.create_sheet("전체데이터")

    # 실제 수집 시작 년월 산출
    if all_rows:
        yms = sorted((int(r["년"]),int(r["월"])) for r in all_rows)
        sy, sm = yms[0]
    else:
        sy, sm = start_year, start_month

    _overview(ws_ov,  all_rows, sy, sm, run_at)
    _city_log(ws_city, all_rows, run_at, is_update=is_upd)
    _missing_display(ws_md, all_rows, sy, sm)
    _empty_values(ws_ev, all_rows)
    _all_data(ws_all, all_rows)

    summary_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(summary_path)
    return run_at


# ══════════════════════════════════════════════════════════════
# 스크래퍼 코어
# ══════════════════════════════════════════════════════════════
class KitaHtmlScraper:

    UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
          "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
    SEA_PROG = "eyJwZ21JZCI6Ijc3MzciLCJ1cHBlclBnbUlkIjoiNzY3NiIsInBnbU5hbWUiOiLH2LvzoaTH17D4IML8sO2/7sDTIMG2yLgiLCJ1cmwiOiIvc2hpcHBlcnMvbG9naXN0aWNzRmFyZS9sb2dpc3RpY3NTZWFGYXJlLmRvIiwibWVudVNldElkIjoiMjQ4MiIsIm1lbnVEZXB0aCI6IjMiLCJ0b3BNZW51SWQiOiIzIiwidG9wTWVudU5hbWUiOiLB9r/4oaS75773IiwibWJyRXh1c1luIjoiTiIsImFjY2VzQXV0aFVzZVluIjoiTiIsImduckFjY2VzUHNibFluIjoiWSIsIm1lbnVUeXBlQ2QiOiIxMCIsInRoZW1lIjoiMyIsInBnbVRpdGwiOiIiLCJwZ21EZXNjciI6IiJ9"
    AIR_PROG = "eyJwZ21JZCI6Ijc3MzciLCJ1cHBlclBnbUlkIjoiNzY3NiIsInBnbU5hbWUiOiLH2LvzoaTH17D4IML8sO2/7sDTIMG2yLgiLCJ1cmwiOiIvc2hpcHBlcnMvbG9naXN0aWNzRmFyZS9sb2dpc3RpY3NBaXJGYXJlLmRvIiwibWVudVNldElkIjoiMjQ4MiIsIm1lbnVEZXB0aCI6IjMiLCJ0b3BNZW51SWQiOiIzIiwidG9wTWVudU5hbWUiOiLB9r/4oaS75773IiwibWJyRXh1c1luIjoiTiIsImFjY2VzQXV0aFVzZVluIjoiTiIsImduckFjY2VzUHNibFluIjoiWSIsIm1lbnVUeXBlQ2QiOiIxMCIsInRoZW1lIjoiMyIsInBnbVRpdGwiOiIiLCJwZ21EZXNjciI6IiJ9"

    def __init__(self, start_year, start_month):
        self.start_year  = start_year
        self.start_month = start_month
        self.today       = datetime.today().strftime("%Y-%m-%d")
        self._stop       = False
        self._session    = requests.Session()
        self._sea_cookie = ""
        self._air_cookie = ""

    def interrupt(self):
        log.warning("⚠️  Ctrl+C → 현재 도시 완료 후 안전 종료...")
        self._stop = True

    # ── STEP1: 세션 쿠키 획득 ────────────────────────────────
    def acquire_session(self):
        """playwright로 1회 접속하여 JSESSIONID 획득"""
        log.info("세션 쿠키 획득 중 (playwright, 약 30초)...")
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            ctx     = browser.new_context(user_agent=self.UA)
            page    = ctx.new_page()

            page.goto(SEA_PAGE, wait_until="networkidle", timeout=30000)
            page.wait_for_timeout(2000)
            sea_cookies = {c["name"]: c["value"] for c in ctx.cookies()}

            page.goto(AIR_PAGE, wait_until="networkidle", timeout=30000)
            page.wait_for_timeout(2000)
            air_cookies = {c["name"]: c["value"] for c in ctx.cookies()}

            browser.close()

        self._sea_cookie = (f"JSESSIONID_KITA={sea_cookies.get('JSESSIONID_KITA','')}; "
                            f"WMONID={sea_cookies.get('WMONID','')}")
        self._air_cookie = (f"JSESSIONID_KITA={air_cookies.get('JSESSIONID_KITA','')}; "
                            f"WMONID={air_cookies.get('WMONID','')}")
        log.info(f"  해상 세션: {sea_cookies.get('JSESSIONID_KITA','')[:20]}...")
        log.info(f"  항공 세션: {air_cookies.get('JSESSIONID_KITA','')[:20]}...")

    # ── 도시 코드 목록 조회 ───────────────────────────────────
    def get_cities(self, freight_type, region_code):
        """logisticsSeaCodeView.do / logisticsAirCodeView.do → 도시 목록"""
        api_url = SEA_CODE_API if freight_type == "sea" else AIR_CODE_API
        cookie  = self._sea_cookie if freight_type == "sea" else self._air_cookie
        hdrs = {
            "User-Agent":       self.UA,
            "Content-Type":     "application/json",
            "X-Requested-With": "XMLHttpRequest",
            "ajax":             "TRUE",
            "Cookie":           cookie,
            "Referer":          SEA_PAGE if freight_type=="sea" else AIR_PAGE,
            "pageprograminfo":  self.SEA_PROG if freight_type=="sea" else self.AIR_PROG,
            "pageauthinfo":     "bnVsbA==",
        }
        try:
            resp = self._session.post(api_url, json={"code": region_code},
                                      headers=hdrs, timeout=10)
            rows = resp.json().get("detailView", [])
            cities = [{"value": r["code"], "text": r.get("codeNm",""),
                       "text_en": r.get("codeEn","")}
                      for r in rows if r.get("code")]
            log.info(f"  권역 {region_code}: 도시 {len(cities)}개")
            return cities
        except Exception as e:
            log.error(f"  도시 목록 조회 실패 ({region_code}): {e}")
            return []

    # ── HTML POST → 테이블 파싱 ───────────────────────────────
    def fetch_page(self, freight_type, start_code, region_code, city_code,
                   region_name, sy, sm, ey, em, page_num=1):
        """requests.post → HTML → BeautifulSoup 파싱 → 행 리스트"""
        page_url = SEA_PAGE if freight_type == "sea" else AIR_PAGE
        cookie   = self._sea_cookie if freight_type == "sea" else self._air_cookie
        hdrs = {
            "User-Agent":    self.UA,
            "Content-Type":  "application/x-www-form-urlencoded",
            "Referer":       page_url,
            "Cookie":        cookie,
            "Accept":        "text/html,application/xhtml+xml,*/*",
            "Accept-Language": "ko-KR,ko;q=0.9",
        }
        form = {
            "searchStart":       start_code,
            "searchEndMainLand": region_code,
            "searchEnd":         city_code,
            "dayChange":         "B",
            "searchYear":        sy,
            "searchMonth":       sm,
            "searchYear2":       ey,
            "searchMonth2":      em,
            "pageIndex":         str(page_num),
        }
        resp = self._session.post(page_url, data=form, headers=hdrs, timeout=20)
        resp.raise_for_status()
        return self._parse_html(resp.text, freight_type, region_name)

    def _parse_html(self, html, freight_type, region_name=""):
        """HTML에서 logistics-tb 테이블 파싱"""
        soup  = BeautifulSoup(html, "html.parser")
        table = soup.find("table", class_="logistics-tb")
        if not table:
            return [], 0

        # 전체 페이지 수 파악
        total_pages = 1
        paging = soup.find(class_=re.compile(r"paging|pagination"))
        if paging:
            nums = [int(a.get_text(strip=True))
                    for a in paging.find_all("a")
                    if a.get_text(strip=True).isdigit()]
            if nums:
                total_pages = max(nums)

        rows = []
        last_dep = last_arr = ""
        tbody = table.find("tbody")
        if not tbody:
            return rows, total_pages

        for tr in tbody.find_all("tr"):
            tds = tr.find_all("td")
            if not tds:
                continue

            # 텍스트 (출발/도착/년/월 확인용)
            texts = [td.get_text(separator=" ", strip=True) for td in tds]
            if not any(texts):
                continue
            if len(texts) == 1 and "없습니다" in texts[0]:
                continue

            # rowspan 채우기 (출발·도착은 텍스트로)
            dep = texts[0] if texts[0] else last_dep
            arr = texts[1] if len(texts) > 1 and texts[1] else last_arr
            if texts[0]:              last_dep = dep
            if len(texts) > 1 and texts[1]: last_arr = arr

            try:
                # 년도 (float "2025.0" 도 처리)
                raw_year = texts[2].strip() if len(texts) > 2 else ""
                try:
                    year = str(int(float(raw_year))) if raw_year else ""
                except (ValueError, TypeError):
                    year = raw_year
                raw_mon = texts[3].strip() if len(texts) > 3 else ""
                # 월 두 자리 포맷 (float 형태 "3.0" 도 처리)
                try:
                    mon = f"{int(float(raw_mon)):02d}" if raw_mon else ""
                except (ValueError, TypeError):
                    mon = raw_mon

                # [버그 수정] 깨진 행 필터링:
                # KITA HTML 페이지 구분선(tr)이 파싱돼 "/-" 같은 잔재가 들어오는 경우
                # 년/월이 없거나 출발도시가 정상값이 아닌 행 제거
                if not year or not mon or not year.isdigit():
                    continue
                if dep and dep not in ("Busan", "Incheon") and last_dep not in ("Busan", "Incheon"):
                    continue

                if freight_type == "sea":
                    # 운임 td를 태그째로 전달 → span(편차) 제거 후 순수 운임값만 추출
                    teu = clean_num_from_td(tds[4] if len(tds) > 4 else None, "TEU")
                    feu = clean_num_from_td(tds[5] if len(tds) > 5 else None, "FEU")
                    row = {
                        "분류": "해상", "출발(도시)": dep, "도착(국가)": region_name,
                        "도착(도시)": arr, "년": str(year), "월": mon,
                        "TEU": teu, "FEU": feu, "단위(USD-해상)": "USD",
                        "100kg": "", "300kg": "", "500kg": "", "단위(USD-항공)": "",
                        "업데이트 일자": self.today,
                    }
                else:
                    k100 = clean_num_from_td(tds[4] if len(tds) > 4 else None, "100kg")
                    k300 = clean_num_from_td(tds[5] if len(tds) > 5 else None, "300kg")
                    k500 = clean_num_from_td(tds[6] if len(tds) > 6 else None, "500kg")
                    row = {
                        "분류": "항공", "출발(도시)": dep, "도착(국가)": region_name,
                        "도착(도시)": arr, "년": str(year), "월": mon,
                        "TEU": "", "FEU": "", "단위(USD-해상)": "",
                        "100kg": k100, "300kg": k300, "500kg": k500,
                        "단위(USD-항공)": "USD",
                        "업데이트 일자": self.today,
                    }
                rows.append(row)
            except Exception:
                continue

        return rows, total_pages

    def fetch_all_pages(self, freight_type, start_code, region_code, city_code,
                        region_name, sy, sm, ey, em):
        """전체 페이지 수집"""
        all_rows = []
        for attempt in range(3):
            try:
                rows, total = self.fetch_page(freight_type, start_code, region_code,
                                               city_code, region_name, sy, sm, ey, em, page_num=1)
                all_rows.extend(rows)
                log.info(f"    페이지 1/{total}: {len(rows)}행")

                for pn in range(2, total + 1):
                    rows_p, _ = self.fetch_page(freight_type, start_code, region_code,
                                                 city_code, region_name, sy, sm, ey, em, page_num=pn)
                    all_rows.extend(rows_p)
                    log.info(f"    페이지 {pn}/{total}: {len(rows_p)}행")
                    time.sleep(0.3)
                break
            except Exception as e:
                if attempt < 2:
                    log.warning(f"    재시도 {attempt+1}/3: {e}")
                    time.sleep(2)
                    # 세션 갱신
                    self.acquire_session()
                else:
                    raise

        return all_rows

    # ── 도시 1개 처리 ────────────────────────────────────────
    def process_city(self, freight_type, region_code, region_name,
                     city_code, city_name, sy, sm, ey, em):
        start_code = SEA_START if freight_type == "sea" else AIR_START
        log.info(f"  ▷ {city_name} ({city_code})")
        for attempt in range(3):
            try:
                rows = self.fetch_all_pages(
                    freight_type, start_code, region_code,
                    city_code, region_name, sy, sm, ey, em
                )
                save_progress(rows, freight_type, region_code, city_code)
                remove_failed(freight_type, city_code)
                log.info(f"    → {len(rows)}행 ✅")
                return rows, True
            except Exception as e:
                log.error(f"    시도 {attempt+1}/3 실패: {e}")
                if attempt < 2:
                    time.sleep(3)

        log.error(f"    ❌ {city_name} 최종 실패")
        append_failed({
            "freight_type": freight_type, "region_code": region_code,
            "region_name":  region_name,  "city_code":   city_code,
            "city_name":    city_name,    "start_year":  self.start_year,
            "start_month":  self.start_month,
            "failed_at":    datetime.now().isoformat(),
        })
        return [], False

    # ── 전체 수집 ────────────────────────────────────────────
    def scrape(self, retry_only=False, skip_done=False):
        ey = str(date.today().year)
        em = str(date.today().month)
        sy = str(self.start_year)
        sm = str(self.start_month)
        log.info(f"수집 기간: {sy}년 {sm}월 ~ {ey}년 {em}월")

        done_cities  = get_done_cities() if skip_done else set()
        total_rows   = 0
        freight_list = [("sea", SEA_REGIONS), ("air", AIR_REGIONS)]

        if retry_only:
            failed = load_failed()
            if not failed:
                log.info("재수집 항목 없음")
                return
            log.info(f"재수집 대상: {len(failed)}개")
            for entry in failed:
                if self._stop: break
                rows, _ = self.process_city(
                    entry["freight_type"], entry["region_code"], entry["region_name"],
                    entry["city_code"], entry["city_name"], sy, sm, ey, em
                )
                total_rows += len(rows)
                time.sleep(0.5)
        else:
            for freight_type, regions in freight_list:
                if self._stop: break
                label = "해상" if freight_type == "sea" else "항공"
                log.info(f"\n{'='*50}\n[{label}] 수집 시작\n{'='*50}")

                for region_code, region_name in regions.items():
                    if self._stop: break
                    log.info(f"\n▶ {label} / {region_name} ({region_code})")

                    cities = self.get_cities(freight_type, region_code)
                    if not cities:
                        log.warning("  도시 없음 - 건너뜀")
                        continue

                    for city in cities:
                        if self._stop: break
                        if skip_done and (freight_type, city["value"]) in done_cities:
                            log.info(f"  ▷ {city['text']} - 이미 수집됨, 건너뜀")
                            continue
                        rows, _ = self.process_city(
                            freight_type, region_code, region_name,
                            city["value"], city["text"], sy, sm, ey, em
                        )
                        total_rows += len(rows)
                        time.sleep(0.5)   # 서버 부하 방지

        if self._stop:
            log.warning("⚠️  중단으로 종료")
        log.info(f"\n수집 완료: {total_rows}행")


# ── 인수 파싱 ─────────────────────────────────────────────────
def parse_args():
    sy, sm = DEFAULT_START_YEAR, DEFAULT_START_MONTH
    schedule = retry = skip_done = False
    args = sys.argv[1:]
    i = 0
    while i < len(args):
        if   args[i] == "--from"       and i+1 < len(args): sy = int(args[i+1]); i += 2
        elif args[i] == "--from-month" and i+1 < len(args): sm = int(args[i+1]); i += 2
        elif args[i] == "--schedule":  schedule  = True; i += 1
        elif args[i] == "--retry":     retry     = True; i += 1
        elif args[i] == "--skip-done": skip_done = True; i += 1
        else: i += 1
    return sy, sm, schedule, retry, skip_done


# ── 실행 ──────────────────────────────────────────────────────
def run_once(start_year=DEFAULT_START_YEAR, start_month=DEFAULT_START_MONTH,
             retry_only=False, skip_done=False):
    log.info(f"\n{'='*60}")
    log.info(f"KITA 운임 수집 시작 (HTML POST 방식): {datetime.now():%Y-%m-%d %H:%M:%S}")
    log.info(f"기간: {start_year}년 {start_month}월~ | 모드: {'재수집' if retry_only else '전체'}")
    log.info(f"{'='*60}")

    scraper = KitaHtmlScraper(start_year, start_month)
    signal.signal(signal.SIGINT, lambda s, f: scraper.interrupt())

    scraper.acquire_session()
    scraper.scrape(retry_only=retry_only, skip_done=skip_done)

    log.info("\n최종 CSV 병합 중...")
    merge_progress_to_csv()

    failed = load_failed()
    if failed:
        log.warning(f"\n⚠️  미수집 {len(failed)}개:")
        for f in failed:
            log.warning(f"   [{f['freight_type']}] {f['region_name']} / {f['city_name']}")
        log.warning("→ 재수집: python kita_api_scraper.py --retry")

    log.info(f"\n완료: {datetime.now():%Y-%m-%d %H:%M:%S}")


def run_scheduled(start_year, start_month):
    try:
        from apscheduler.schedulers.blocking import BlockingScheduler
        from apscheduler.triggers.cron import CronTrigger
        import functools
    except ImportError:
        log.error("pip install apscheduler 필요")
        sys.exit(1)

    sched = BlockingScheduler(timezone="Asia/Seoul")
    sched.add_job(
        functools.partial(run_once, start_year, start_month),
        CronTrigger(day=9, hour=8, minute=0, timezone="Asia/Seoul"),
        id="kita_monthly", replace_existing=True,
    )
    log.info("스케줄러: 매월 9일 08:00 KST  |  종료: Ctrl+C")
    sched.start()


if __name__ == "__main__":
    sy, sm, schedule, retry, skip_done = parse_args()
    if schedule:
        run_scheduled(sy, sm)
    else:
        run_once(sy, sm, retry_only=retry, skip_done=skip_done)
